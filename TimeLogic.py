import openpyxl
from datetime import datetime


wb = openpyxl.load_workbook('MRI_data1.xlsx')
#print(wb.sheetnames)
sheet = wb['Sheet2']
sheet['L1'] = 'Time Elapsed'
row_count = sheet.max_row
final_data = []
for i in range(2,row_count+1):
     data1 = sheet.cell(row=i, column=5).value
     data2 = sheet.cell(row=i, column=6).value
     difference = data1 - data2
#to save the generated data to new column
     sheet.cell(row=i, column=12).value = divmod(difference.days * 86400 + difference.seconds, 60)[0]

#To save the excel sheet
wb.save('MRI_data1.xlsx')
